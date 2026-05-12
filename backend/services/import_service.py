"""
Koopilot Veri Iceri Aktarma Servisi
────────────────────────────────────
Excel (.xlsx) ve CSV dosyalarindan veritabanina toplu veri aktarimi.
Desteklenen veri tipleri: urunler, siparisler, tedarikciler
"""

import csv
import io
from openpyxl import load_workbook
from database import SessionLocal
from models import Product, Order, OrderItem, Supplier


# ── Kolon esleme haritasi (Turkce baslik -> model alani) ──────────────────────

PRODUCT_COLUMN_MAP = {
    "id": "id", "kod": "id", "urun_kodu": "id", "ürün kodu": "id",
    "ad": "name", "urun_adi": "name", "ürün adı": "name", "ürün": "name", "name": "name",
    "kategori": "category", "category": "category",
    "stok": "stock_quantity", "stok_miktari": "stock_quantity", "stok miktarı": "stock_quantity", "stock": "stock_quantity", "miktar": "stock_quantity",
    "esik": "low_threshold", "eşik": "low_threshold", "kritik_esik": "low_threshold", "threshold": "low_threshold",
    "fiyat": "unit_price", "birim_fiyat": "unit_price", "birim fiyat": "unit_price", "price": "unit_price",
    "birim": "unit", "unit": "unit",
    "tedarikci_id": "supplier_id", "tedarikçi": "supplier_id", "supplier": "supplier_id",
}

ORDER_COLUMN_MAP = {
    "id": "id", "siparis_no": "id", "sipariş no": "id", "siparis_id": "id",
    "musteri": "customer_name", "müşteri": "customer_name", "musteri_adi": "customer_name", "müşteri adı": "customer_name", "customer": "customer_name",
    "telefon": "customer_phone", "tel": "customer_phone", "phone": "customer_phone",
    "email": "customer_email", "e-posta": "customer_email",
    "durum": "status", "status": "status",
    "tutar": "total_amount", "toplam": "total_amount", "toplam tutar": "total_amount", "amount": "total_amount",
    "kargo_takip": "tracking_number", "takip_no": "tracking_number", "tracking": "tracking_number",
    "kargo_firma": "cargo_company", "kargo": "cargo_company",
    "not": "notes", "notlar": "notes", "notes": "notes",
}

SUPPLIER_COLUMN_MAP = {
    "id": "id", "kod": "id", "tedarikci_kodu": "id",
    "ad": "name", "firma": "name", "tedarikci_adi": "name", "tedarikçi adı": "name", "name": "name",
    "email": "email", "e-posta": "email",
    "telefon": "phone", "tel": "phone", "phone": "phone",
}


def _normalize_header(h):
    """Baslik metnini normalize et."""
    return h.strip().lower().replace("İ", "i").replace("ı", "i").replace("ş", "s").replace("ç", "c").replace("ö", "o").replace("ü", "u")


def _detect_data_type(headers):
    """Kolon basliklarindan veri tipini otomatik tespit et."""
    h_set = set(headers)
    product_matches = sum(1 for h in h_set if h in PRODUCT_COLUMN_MAP)
    order_matches = sum(1 for h in h_set if h in ORDER_COLUMN_MAP)
    supplier_matches = sum(1 for h in h_set if h in SUPPLIER_COLUMN_MAP)

    if product_matches >= order_matches and product_matches >= supplier_matches:
        return "products"
    elif order_matches >= supplier_matches:
        return "orders"
    else:
        return "suppliers"


def _map_headers(headers, column_map):
    """Baslik satirini model alanlarina esle."""
    mapped = {}
    for i, h in enumerate(headers):
        norm = _normalize_header(h)
        if norm in column_map:
            mapped[i] = column_map[norm]
        elif h.strip().lower() in column_map:
            mapped[i] = column_map[h.strip().lower()]
    return mapped


def _parse_excel(file_bytes):
    """Excel dosyasini satirlara cevir."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(c or "").strip() for c in rows[0]]
    data = []
    for row in rows[1:]:
        if any(c is not None for c in row):
            data.append([c for c in row])
    return headers, data


def _parse_csv(file_bytes):
    """CSV dosyasini satirlara cevir."""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    headers = [h.strip() for h in rows[0]]
    data = rows[1:]
    return headers, data


def _safe_int(val, default=0):
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _safe_float(val, default=0.0):
    try:
        return float(str(val).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return default


def import_file(file_bytes: bytes, filename: str, data_type: str = None) -> dict:
    """
    Ana import fonksiyonu.
    
    Args:
        file_bytes: Dosya icerigi
        filename: Dosya adi (.xlsx veya .csv)
        data_type: "products", "orders", "suppliers" veya None (otomatik tespit)
    
    Returns:
        {"success": bool, "imported": int, "skipped": int, "errors": list, "data_type": str}
    """
    ext = filename.rsplit(".", 1)[-1].lower()
    
    if ext in ("xlsx", "xls"):
        headers, data = _parse_excel(file_bytes)
    elif ext == "csv":
        headers, data = _parse_csv(file_bytes)
    else:
        return {"success": False, "imported": 0, "skipped": 0, "errors": ["Desteklenmeyen dosya formati. .xlsx veya .csv kullanin."], "data_type": "unknown"}

    if not headers or not data:
        return {"success": False, "imported": 0, "skipped": 0, "errors": ["Dosya bos veya baslik satiri bulunamadi."], "data_type": "unknown"}

    # Baslik normalizasyonu
    norm_headers = [_normalize_header(h) for h in headers]
    
    # Veri tipi tespiti
    if not data_type:
        data_type = _detect_data_type(norm_headers)

    if data_type == "products":
        return _import_products(norm_headers, data)
    elif data_type == "orders":
        return _import_orders(norm_headers, data)
    elif data_type == "suppliers":
        return _import_suppliers(norm_headers, data)
    else:
        return {"success": False, "imported": 0, "skipped": 0, "errors": [f"Bilinmeyen veri tipi: {data_type}"], "data_type": data_type}


def _import_products(headers, data):
    col_map = _map_headers(headers, PRODUCT_COLUMN_MAP)
    db = SessionLocal()
    imported, skipped = 0, 0
    errors = []

    try:
        for row_idx, row in enumerate(data, start=2):
            try:
                record = {}
                for col_idx, field_name in col_map.items():
                    if col_idx < len(row):
                        record[field_name] = row[col_idx]

                if "name" not in record or not record["name"]:
                    skipped += 1
                    continue

                pid = str(record.get("id", f"PRD-AUTO-{row_idx}")).strip()
                existing = db.query(Product).filter(Product.id == pid).first()

                if existing:
                    existing.name = str(record["name"]).strip()
                    existing.category = str(record.get("category", "genel")).strip()
                    existing.stock_quantity = _safe_int(record.get("stock_quantity", 0))
                    existing.low_threshold = _safe_int(record.get("low_threshold", 10))
                    existing.unit_price = _safe_float(record.get("unit_price", 0))
                    existing.unit = str(record.get("unit", "adet")).strip()
                    if record.get("supplier_id"):
                        existing.supplier_id = str(record["supplier_id"]).strip()
                else:
                    product = Product(
                        id=pid,
                        name=str(record["name"]).strip(),
                        category=str(record.get("category", "genel")).strip(),
                        stock_quantity=_safe_int(record.get("stock_quantity", 0)),
                        low_threshold=_safe_int(record.get("low_threshold", 10)),
                        unit_price=_safe_float(record.get("unit_price", 0)),
                        unit=str(record.get("unit", "adet")).strip(),
                        supplier_id=str(record["supplier_id"]).strip() if record.get("supplier_id") else None,
                    )
                    db.add(product)
                imported += 1
            except Exception as e:
                errors.append(f"Satir {row_idx}: {str(e)[:80]}")
                skipped += 1

        db.commit()
    finally:
        db.close()

    return {"success": True, "imported": imported, "skipped": skipped, "errors": errors, "data_type": "products"}


def _import_orders(headers, data):
    col_map = _map_headers(headers, ORDER_COLUMN_MAP)
    db = SessionLocal()
    imported, skipped = 0, 0
    errors = []

    try:
        for row_idx, row in enumerate(data, start=2):
            try:
                record = {}
                for col_idx, field_name in col_map.items():
                    if col_idx < len(row):
                        record[field_name] = row[col_idx]

                if "customer_name" not in record or not record["customer_name"]:
                    skipped += 1
                    continue

                oid = str(record.get("id", f"ORD-AUTO-{row_idx}")).strip()
                existing = db.query(Order).filter(Order.id == oid).first()

                if existing:
                    existing.customer_name = str(record["customer_name"]).strip()
                    existing.status = str(record.get("status", "beklemede")).strip()
                    existing.total_amount = _safe_float(record.get("total_amount", 0))
                    if record.get("customer_phone"):
                        existing.customer_phone = str(record["customer_phone"]).strip()
                    if record.get("tracking_number"):
                        existing.tracking_number = str(record["tracking_number"]).strip()
                    if record.get("cargo_company"):
                        existing.cargo_company = str(record["cargo_company"]).strip()
                else:
                    order = Order(
                        id=oid,
                        customer_name=str(record["customer_name"]).strip(),
                        customer_phone=str(record.get("customer_phone", "")).strip() or None,
                        customer_email=str(record.get("customer_email", "")).strip() or None,
                        status=str(record.get("status", "beklemede")).strip(),
                        total_amount=_safe_float(record.get("total_amount", 0)),
                        tracking_number=str(record.get("tracking_number", "")).strip() or None,
                        cargo_company=str(record.get("cargo_company", "")).strip() or None,
                        notes=str(record.get("notes", "")).strip() or None,
                    )
                    db.add(order)
                imported += 1
            except Exception as e:
                errors.append(f"Satir {row_idx}: {str(e)[:80]}")
                skipped += 1

        db.commit()
    finally:
        db.close()

    return {"success": True, "imported": imported, "skipped": skipped, "errors": errors, "data_type": "orders"}


def _import_suppliers(headers, data):
    col_map = _map_headers(headers, SUPPLIER_COLUMN_MAP)
    db = SessionLocal()
    imported, skipped = 0, 0
    errors = []

    try:
        for row_idx, row in enumerate(data, start=2):
            try:
                record = {}
                for col_idx, field_name in col_map.items():
                    if col_idx < len(row):
                        record[field_name] = row[col_idx]

                if "name" not in record or not record["name"]:
                    skipped += 1
                    continue

                sid = str(record.get("id", f"SUP-AUTO-{row_idx}")).strip()
                existing = db.query(Supplier).filter(Supplier.id == sid).first()

                if existing:
                    existing.name = str(record["name"]).strip()
                    if record.get("email"):
                        existing.email = str(record["email"]).strip()
                    if record.get("phone"):
                        existing.phone = str(record["phone"]).strip()
                else:
                    supplier = Supplier(
                        id=sid,
                        name=str(record["name"]).strip(),
                        email=str(record.get("email", "")).strip() or None,
                        phone=str(record.get("phone", "")).strip() or None,
                    )
                    db.add(supplier)
                imported += 1
            except Exception as e:
                errors.append(f"Satir {row_idx}: {str(e)[:80]}")
                skipped += 1

        db.commit()
    finally:
        db.close()

    return {"success": True, "imported": imported, "skipped": skipped, "errors": errors, "data_type": "suppliers"}


def preview_file(file_bytes: bytes, filename: str) -> dict:
    """Dosyanin ilk 5 satirini onizleme olarak dondurur."""
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext in ("xlsx", "xls"):
        headers, data = _parse_excel(file_bytes)
    elif ext == "csv":
        headers, data = _parse_csv(file_bytes)
    else:
        return {"headers": [], "rows": [], "detected_type": "unknown", "total_rows": 0}

    norm_headers = [_normalize_header(h) for h in headers]
    detected = _detect_data_type(norm_headers)

    return {
        "headers": headers,
        "rows": [list(r) for r in data[:5]],
        "detected_type": detected,
        "total_rows": len(data),
    }
